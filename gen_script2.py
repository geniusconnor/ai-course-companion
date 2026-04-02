# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "徐糖果·立人物·分镜脚本"

SEG_COLORS = {1: "FFF3CD", 2: "D4EDDA", 3: "CCE5FF", 4: "F8D7DA"}
HEADER_FILL = PatternFill("solid", fgColor="2D2D2D")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
TITLE_FILL  = PatternFill("solid", fgColor="1A1A2E")
TITLE_FONT  = Font(name="微软雅黑", bold=True, color="FFFFFF", size=13)

thin   = Side(style="thin",   color="BBBBBB")
medium = Side(style="medium", color="888888")
THIN_BORDER = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
SEG_BORDER  = Border(left=medium, right=medium, top=medium, bottom=medium)
WRAP   = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

segments = [
    {
        "seq": 1,
        "name": "先感受她「停不下来」",
        "function": (
            "让观众身体上感受到：她是一直在动、停不下来、能量特别满的小孩。"
            "先别解释她是谁。"
        ),
        "duration": "约 15–20 秒\n（整体 50 秒–1 分 10 秒）",
        "shots": [
            ("1-A", "在家骑平衡车"),
            ("1-B", "在家里跑跳的画面\n（后期补：厨房爬墙视频）"),
            ("1-C", "拿到蛋糕时兴奋反应"),
        ],
        "sync_voice": (
            "「我没有安静的时候，只有睡觉。」\n\n"
            "→ 尽量早出，甚至可做开场第一句"
        ),
        "edit_notes": (
            "· 节奏快，多用她在动的镜头\n"
            "· 这句台词尽量早出，可做片头第一句\n"
            "· 不要加解说，靠画面让观众先「感受到」"
        ),
    },
    {
        "seq": 2,
        "name": "进入她的世界",
        "function": (
            "让观众知道：她有一个很明确的世界中心。"
            "不是展示优秀，而是她已经很清楚自己最重要的东西是什么。"
        ),
        "duration": "15–20 秒",
        "shots": [
            ("2-A", "发小天才手表朋友圈"),
            ("2-B", "奖牌盒子（她与盒子互动）"),
            ("2-C", "攀岩墙"),
            ("2-D", "她介绍房间里的书 / 奖牌 / 自己的东西"),
        ],
        "sync_voice": "「我只是偶尔看书，但一直在攀岩。」",
        "edit_notes": (
            "· 奖牌不要停太久，重点是「这些是她生活的一部分」\n"
            "· 优先留「她与奖牌盒子互动」，而非奖牌全景\n"
            "· 整段要有「她在主动带你参观她的宇宙」的感觉"
        ),
    },
    {
        "seq": 3,
        "name": "她有自己的语言系统",
        "function": (
            "把她从「有活力的小孩」升级成有边界、有判断、"
            "有自己表达方式的人。她会自己命名自己，也会拆别人给她的词。"
        ),
        "duration": "15–20 秒",
        "shots": [
            ("3-A", "互贴标签"),
            ("3-B", "妈妈给她贴标签"),
            ("3-C", "她给自己贴标签"),
            ("3-D", "她吐槽「自由旷野是人设词」"),
            ("3-E", "她说不喜欢别人夸她漂亮"),
        ],
        "sync_voice": (
            "推荐顺序：\n"
            "① 妈妈说她：「阳光、自由、有个核心能力」\n"
            "② 她说自己：「傻、抽象、搞笑」\n"
            "③「自由旷野之类的词，就是人设词。」\n"
            "④「我不喜欢别人夸我漂亮，因为我不喜欢别人评价我。」"
        ),
        "edit_notes": (
            "· 这段不要太快\n"
            "· 「人设词」和「不喜欢别人评价我」两句，最好给一点停顿\n"
            "· 允许稍微「锋利」一点，这正是她的魅力\n"
            "· 这段会让人物一下变立体"
        ),
    },
    {
        "seq": 4,
        "name": "身体状态把人物做实",
        "function": (
            "告诉观众：她的「停不下来」不是嘴贫、不是表演感，"
            "是她整个人的真实状态——脑子和身体都在往外长。"
        ),
        "duration": "10–15 秒",
        "shots": [
            ("4-A", "攀岩训练时累到手虚"),
            ("4-B", "表情明显疲惫"),
            ("4-C", "休息一下又立刻恢复"),
            ("4-D", "训练结束后还想再爬一条（可选）"),
        ],
        "sync_voice": (
            "可选旁白：\n"
            "「她不只是嘴上热闹，连身体都像一直在往前扑。」\n\n"
            "→ 也可不用旁白，直接靠镜头"
        ),
        "edit_notes": (
            "· 重点不是她多厉害\n"
            "· 而是她真的会累，但累完还会回来\n"
            "· 这会让前面的「高能量」变得可信"
        ),
    },
]

col_widths = {1: 8, 2: 18, 3: 34, 4: 18, 5: 8, 6: 26, 7: 36, 8: 40}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# 标题行
ws.merge_cells("A1:H1")
ws["A1"] = "徐糖果·立人物  分镜头脚本"
ws["A1"].font = TITLE_FONT
ws["A1"].fill = TITLE_FILL
ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 32

# 表头行
headers = ["段落序号", "段落名称", "段落功能 / 目的", "预估时长",
           "镜头编号", "镜头内容", "同期声 / 旁白", "剪辑提醒"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
ws.row_dimensions[2].height = 28

# 数据行
row = 3
for seg in segments:
    seg_fill = PatternFill("solid", fgColor=SEG_COLORS[seg["seq"]])
    n_shots = len(seg["shots"])
    start_row = row

    for shot_no, shot_desc in seg["shots"]:
        for col in range(1, 9):
            c = ws.cell(row=row, column=col)
            c.fill = seg_fill
            c.border = THIN_BORDER
            c.alignment = WRAP
        ws.cell(row=row, column=5, value=shot_no).alignment = CENTER
        ws.cell(row=row, column=6, value=shot_desc).alignment = WRAP
        row += 1

    end_row = row - 1

    merge_data = [
        (1, "第" + str(seg["seq"]) + "段"),
        (2, seg["name"]),
        (3, seg["function"]),
        (4, seg["duration"]),
        (7, seg["sync_voice"]),
        (8, seg["edit_notes"]),
    ]
    for col, val in merge_data:
        if n_shots > 1:
            ws.merge_cells(
                start_row=start_row, start_column=col,
                end_row=end_row,     end_column=col
            )
        cell = ws.cell(row=start_row, column=col, value=val)
        cell.fill = seg_fill
        cell.border = SEG_BORDER
        cell.alignment = WRAP
        if col == 1:
            cell.alignment = CENTER
            cell.font = Font(name="微软雅黑", bold=True, size=11)
        elif col == 2:
            cell.font = Font(name="微软雅黑", bold=True, size=10)

    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = 72

# 整体逻辑行
logic_fill = PatternFill("solid", fgColor="EAE8F0")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
logic_text = (
    "【整体逻辑】"
    "先感受到她停不下来 → 再看到她有自己的世界 → "
    "再发现她很有自己 → 最后用身体状态把这一切坐实  |  "
    "不按时间顺序，按观众认识她的顺序剪  |  "
    "目标：先让观众喜欢她、记住她，而不是先讲清楚她会攀岩"
)
cell = ws.cell(row=row, column=1, value=logic_text)
cell.fill = logic_fill
cell.font = Font(name="微软雅黑", bold=True, size=10, color="3A3A6A")
cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
cell.border = SEG_BORDER
ws.row_dimensions[row].height = 48

ws.freeze_panes = "A3"
out_path = r"e:\伴读书童\徐糖果_立人物_分镜脚本.xlsx"
wb.save(out_path)
print("saved:", out_path)
